from __future__ import annotations

import io
import unicodedata
from dataclasses import dataclass, field
from datetime import date, datetime

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.exceptions import NotFoundError
from app.models.cliente import Cliente
from app.repositories import cliente_repository
from app.schemas.cliente import ClienteCreate, ClienteUpdate

# Maps normalized column names to model field names
_COLUMN_MAP: dict[str, str] = {
    "codigo": "codigo",
    "código": "codigo",
    "cod": "codigo",
    "nome": "nome",
    "cliente": "nome",
    "razao social": "nome",
    "razão social": "nome",
    "contrato": "contrato",
    "num contrato": "contrato",
    "numero contrato": "contrato",
    "número contrato": "contrato",
    "data do pedido": "data_do_pedido",
    "data pedido": "data_do_pedido",
    "dt pedido": "data_do_pedido",
    "datapedido": "data_do_pedido",
    "data_do_pedido": "data_do_pedido",
    "data da entrega": "data_da_entrega",
    "data entrega": "data_da_entrega",
    "dt entrega": "data_da_entrega",
    "previsao entrega": "data_da_entrega",
    "previsão entrega": "data_da_entrega",
    "data_da_entrega": "data_da_entrega",
    "descricao do pedido": "descricao_do_pedido",
    "descrição do pedido": "descricao_do_pedido",
    "descricao_do_pedido": "descricao_do_pedido",
    "descricao": "descricao_do_pedido",
    "descrição": "descricao_do_pedido",
    "pedido": "descricao_do_pedido",
    "cnpj": "cnpj",
    "cpf": "cpf",
    "ie": "ie",
    "insc estadual": "ie",
    "inscricao estadual": "ie",
    "inscrição estadual": "ie",
    "rg": "rg",
    "endereco": "endereco",
    "endereço": "endereco",
    "logradouro": "endereco",
    "bairro": "bairro",
    "cidade": "cidade",
    "municipio": "cidade",
    "município": "cidade",
    "estado": "estado",
    "uf": "estado",
    "cep": "cep",
    "celular": "celular",
    "telefone": "celular",
    "fone": "celular",
    "tel": "celular",
    "whatsapp": "celular",
    "email": "email",
    "e-mail": "email",
    "responsavel legal": "responsavel_legal",
    "responsável legal": "responsavel_legal",
    "responsavel_legal": "responsavel_legal",
    "responsavel": "responsavel_legal",
    "responsável": "responsavel_legal",
    "cpf responsavel legal": "cpf_responsavel_legal",
    "cpf responsavel": "cpf_responsavel_legal",
    "cpf do responsavel": "cpf_responsavel_legal",
    "cpf_responsavel_legal": "cpf_responsavel_legal",
}


def _normalize(text: str) -> str:
    normalized = unicodedata.normalize("NFD", text.lower().strip())
    return "".join(c for c in normalized if unicodedata.category(c) != "Mn")


def _parse_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            return from_excel(int(value)).date()
        except Exception:
            return None
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _find_header_row(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
) -> tuple[int, dict[int, str]] | None:
    """Scan up to 20 rows for one containing recognizable column headers."""
    for row_idx, row in enumerate(sheet.iter_rows(max_row=20), start=1):
        mapping: dict[int, str] = {}
        for col_idx, cell in enumerate(row):
            raw = str(cell.value or "").strip()
            if not raw:
                continue
            key = _normalize(raw)
            if key in _COLUMN_MAP:
                mapping[col_idx] = _COLUMN_MAP[key]
        if len(mapping) >= 2:
            return row_idx, mapping
    return None


@dataclass
class ImportResult:
    imported: int = 0
    errors: list[str] = field(default_factory=list)


async def listar(db: AsyncSession, search: str | None = None) -> list[Cliente]:
    return await cliente_repository.get_all(db, search=search)


async def buscar(db: AsyncSession, cliente_id: str) -> Cliente:
    cliente = await cliente_repository.get_by_id(db, cliente_id)
    if not cliente:
        raise NotFoundError("Cliente não encontrado")
    return cliente


async def criar(db: AsyncSession, data: ClienteCreate) -> Cliente:
    return await cliente_repository.create(db, data)


async def atualizar(db: AsyncSession, cliente_id: str, data: ClienteUpdate) -> Cliente:
    cliente = await buscar(db, cliente_id)
    return await cliente_repository.update(db, cliente, data)


async def deletar(db: AsyncSession, cliente_id: str) -> None:
    cliente = await buscar(db, cliente_id)
    await cliente_repository.delete(db, cliente)


async def importar_xlsx(db: AsyncSession, conteudo: bytes) -> ImportResult:
    result = ImportResult()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True)
    except Exception as exc:
        result.errors.append(f"Não foi possível abrir o arquivo: {exc}")
        return result

    sheet = wb.active

    header_info = _find_header_row(sheet)
    if header_info is None:
        result.errors.append(
            "Nenhuma linha de cabeçalho reconhecida. "
            "Verifique se o arquivo contém as colunas esperadas."
        )
        return result

    header_row, col_map = header_info
    to_import: list[ClienteCreate] = []

    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=header_row + 1), start=header_row + 1
    ):
        values: dict[str, object] = {}
        for col_idx, cell in enumerate(row):
            if col_idx in col_map:
                values[col_map[col_idx]] = cell.value

        # skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in values.values()):
            continue

        nome = str(values.get("nome") or "").strip()
        if not nome:
            result.errors.append(f"Linha {row_idx}: campo 'nome' obrigatório ausente.")
            continue

        try:
            data = ClienteCreate(
                codigo=str(values.get("codigo") or "").strip() or None,
                nome=nome,
                contrato=str(values.get("contrato") or "").strip() or None,
                data_do_pedido=_parse_date(values.get("data_do_pedido")),
                data_da_entrega=_parse_date(values.get("data_da_entrega")),
                descricao_do_pedido=str(values.get("descricao_do_pedido") or "").strip()
                or None,
                cnpj=str(values.get("cnpj") or "").strip() or None,
                cpf=str(values.get("cpf") or "").strip() or None,
                ie=str(values.get("ie") or "").strip() or None,
                rg=str(values.get("rg") or "").strip() or None,
                endereco=str(values.get("endereco") or "").strip() or None,
                bairro=str(values.get("bairro") or "").strip() or None,
                cidade=str(values.get("cidade") or "").strip() or None,
                estado=str(values.get("estado") or "").strip() or None,
                cep=str(values.get("cep") or "").strip() or None,
                celular=str(values.get("celular") or "").strip() or None,
                email=str(values.get("email") or "").strip() or None,
                responsavel_legal=str(values.get("responsavel_legal") or "").strip()
                or None,
                cpf_responsavel_legal=str(
                    values.get("cpf_responsavel_legal") or ""
                ).strip()
                or None,
            )
            to_import.append(data)
        except Exception as exc:
            result.errors.append(f"Linha {row_idx}: {exc}")

    if to_import:
        await cliente_repository.create_many(db, to_import)
        result.imported = len(to_import)

    return result
